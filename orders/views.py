from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from courses.models import Pack
from .models import Commande, CommandePack, AccesPack  # ← Ajoutez cette ligne
from django.conf import settings
from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import HttpResponse
from django.utils import timezone
from dashboard.models import Parametres
from django.db import models  # Pour les Q objects


import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import Commande

def export_excel(request):
    commandes = Commande.objects.filter(statut='activee').order_by('-date_commande')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Abonnements"
    
    # En-têtes
    headers = ['Date', 'Client', 'Email', 'Téléphone', 'Packs', 'Montant', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Données
    for row, cmd in enumerate(commandes, 2):
        packs = ", ".join([cp.pack.nom for cp in cmd.packs_commandes.all()])
        ws.cell(row=row, column=1, value=cmd.date_commande.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=f"{cmd.utilisateur.first_name} {cmd.utilisateur.last_name}")
        ws.cell(row=row, column=3, value=cmd.utilisateur.email)
        ws.cell(row=row, column=4, value=cmd.utilisateur.telephone or '')
        ws.cell(row=row, column=5, value=packs)
        ws.cell(row=row, column=6, value=float(cmd.montant_total))
        ws.cell(row=row, column=7, value=cmd.get_statut_display())
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="abonnements.xlsx"'
    wb.save(response)
    return response

def envoyer_notification_admin(commande):
    sujet = f"Nouveau reçu pour la commande #{commande.id}"
    message = f"""
    Bonjour,

    Un nouveau reçu a été uploadé pour la commande n°{commande.id}.

    Client : {commande.utilisateur.username} ({commande.utilisateur.email})
    Montant : {commande.montant_total} DA
    Date : {commande.date_commande}

    Connectez-vous à l'admin pour vérifier et activer la commande :
    http://127.0.0.1:8000/admin/orders/commande/{commande.id}/change/
    """
    
    send_mail(
        sujet,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [settings.ADMIN_EMAIL],
        fail_silently=False,
    )

# Panier en session (simple pour commencer)
@login_required
def add_to_cart(request, pack_id):
    pack = get_object_or_404(Pack, id=pack_id)
    cart = request.session.get('cart', [])
    cart.append(pack_id)
    request.session['cart'] = cart
    messages.success(request, f"{pack.nom} ajouté au panier")
    return redirect('courses:pack_detail', pack_id=pack_id)

@login_required
def cart_view(request):
    cart_ids = request.session.get('cart', [])
    packs = Pack.objects.filter(id__in=cart_ids)
    total = sum(pack.prix for pack in packs)
    return render(request, 'orders/cart.html', {'packs': packs, 'total': total})

@login_required
def checkout(request):
    if request.method == 'POST':
        cart_ids = request.session.get('cart', [])
        packs = Pack.objects.filter(id__in=cart_ids)
        total = sum(pack.prix for pack in packs)
        
        commande = Commande.objects.create(
            utilisateur=request.user,
            montant_total=total,
            statut='en_attente'
        )
        
        for pack in packs:
            CommandePack.objects.create(
                commande=commande,
                pack=pack,
                prix_au_moment_achat=pack.prix
            )
        
        request.session['cart'] = []
        
        if request.FILES.get('recu'):
            commande.recu = request.FILES['recu']
            commande.save()
            # Envoyer la notification à l'admin
            envoyer_notification_admin(commande)
        
        messages.success(request, "Votre commande a été enregistrée.")
        return redirect('orders:mes_commandes')
    
    return redirect('orders:cart')

def my_videos(request):
    params = Parametres.get()
    if params.acces_bloques:
        messages.error(request, "Accès bloqué : fin d'année scolaire")
        return redirect('accueil')
    acces_packs = AccesPack.objects.filter(
        utilisateur=request.user
    ).filter(models.Q(date_expiration__isnull=True) | models.Q(date_expiration__gte=timezone.now().date()))
    return render(request, 'orders/my_videos.html', {'acces_packs': acces_packs})

@login_required
def mes_commandes(request):
    commandes = Commande.objects.filter(utilisateur=request.user).order_by('-date_commande')
    return render(request, 'orders/mes_commandes.html', {'commandes': commandes})

@login_required
def upload_recu(request, commande_id):
    commande = get_object_or_404(Commande, id=commande_id, utilisateur=request.user)
    if request.method == 'POST' and request.FILES.get('recu'):
        commande.recu = request.FILES['recu']
        commande.save()
        # Envoyer la notification à l'admin
        envoyer_notification_admin(commande)
        messages.success(request, "Reçu uploadé avec succès. En attente de validation par l'admin.")
    return redirect('orders:mes_commandes')

@login_required
def remove_from_cart(request, pack_id):
    cart = request.session.get('cart', [])
    if pack_id in cart:
        cart.remove(pack_id)
        request.session['cart'] = cart
        messages.success(request, "Article retiré du panier")
    else:
        messages.error(request, "Article non trouvé dans le panier")
    return redirect('orders:cart')

def test_email(request):
    try:
        send_mail(
            'Test de notification - EduVideo',
            'Bonjour,\n\nVotre configuration Resend fonctionne parfaitement !\n\nCordialement,\nEduVideo',
            settings.DEFAULT_FROM_EMAIL,
            [settings.ADMIN_EMAIL],
            fail_silently=False,
        )
        return HttpResponse("✅ Email envoyé avec succès ! Vérifiez votre boîte de réception.")
    except Exception as e:
        return HttpResponse(f"❌ Erreur : {e}")